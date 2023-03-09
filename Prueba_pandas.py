from openpyxl import load_workbook
import datetime


def Agergar_Linea(data):
    wb = load_workbook(filename = 'BITÁCORA_ACTIVIDADES_SIS_2023.xlsx')
    
    #BITÁCORA
    #NO.	FECHA	NIVEL	ÁREA	ACTIVIDAD	ATENCIÓN	TT	OBSERVACIONES
    #
    ws1 = wb["BITÁCORA"]
    
    maxrow=ws1.max_row
    ultimafila = 1
    for r in range (1, maxrow):
        if not ws1.cell(row = r, column= 1).value is None:
            ultimafila = r + 1
        else: break
    #INSERT INTO `bitacora` (`ID`, `Fecha`,          `Nivel`, `Area`, `Actividad`, `Atencion`, `TT`, `Observaciones`) VALUES ('1', current_timestamp(), 'Planta baja', 'Tecnologias', 'mantenimiento de prueba', 'atencion de algo', 'no se que sea', 'sin observacion')
    
    ws1.cell(row=ultimafila, column=1, value= ultimafila-1)
    ws1.cell(row=ultimafila, column=2, value= datetime.date.today())
    ws1.cell(row=ultimafila, column=3, value= data[0])
    ws1.cell(row=ultimafila, column=4, value= data[1])
    ws1.cell(row=ultimafila, column=5, value= data[2])
    ws1.cell(row=ultimafila, column=6, value= data[3])
    ws1.cell(row=ultimafila, column=7, value= data[4])
    ws1.cell(row=ultimafila, column=8, value= data[5])
    print("insertado correctamente \n")

    #,data[2],      data[3],  data[4],data[5]])

    wb.save("BITÁCORA_ACTIVIDADES_SIS_2023.xlsx")
    wb.close()

